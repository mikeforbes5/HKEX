import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import multiprocessing
import xlrd
import urllib3

http = urllib3.PoolManager()
urllib3.disable_warnings()
workers = multiprocessing.cpu_count() * 2 + 1
path = 'C:\\Users\\Michael.Forbes\\PycharmProjects\\HKEX\\'
writer = pd.ExcelWriter('HKEX.xlsx')
pd.set_option('display.max_columns', None)
url = 'https://www.hkex.com.hk/eng/market/sec_tradinfo/tradarng/tradarng_news/currentmonth/news.htm'
r = requests.get(url)
data = r.text
soup = BeautifulSoup(data, features="lxml")
parsed_table = soup.find_all('table')[2]
data = [['https://www.hkex.com.hk/eng/market/sec_tradinfo/tradarng/tradarng_news/currentmonth/' + td.a['href'] if td.find('a') else
             ''.join(td.stripped_strings)
            for td in row.find_all('td')]
            for row in parsed_table.find_all('tr')]

df = pd.DataFrame(data[1:], columns=data[4])
print(df)
df.to_excel(excel_writer=writer, startrow=1,startcol=1, header=False)
tables = pd.read_html(url, header=0)
assert len(tables) == 3
table = tables[2]
print(table)

table.to_excel(excel_writer=writer, startrow=0, header=True)
writer.save()
book = load_workbook(path + 'HKEX.xlsx')
sheet = book.active
workbook = xlrd.open_workbook(path + 'HKEX.xlsx')
worksheet = workbook.sheet_by_index(0)
rows = worksheet.nrows
x = 1
i = 2
while x < rows:
    url = worksheet.cell_value(x, 5)
    try:
        response = http.request('GET', url)
    except:
        x = x + 1
        i = i + 1
        pass
    else:
        soup = BeautifulSoup(response.data, features="lxml")
        for tag in soup.find_all(class_="hkex-maincontent-header-with-bottom-line"):
            print(tag.get_text())
            sheet.cell(i, 7).value = tag.get_text()
        for tag in soup.find_all("pre"):
            print(tag.get_text())
            sheet.cell(i, 8).value = tag.get_text()
            book.save(path + 'HKEX.xlsx')
        x = x + 1
        i = i + 1






